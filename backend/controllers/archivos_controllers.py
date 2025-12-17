import io
import os
from datetime import datetime
from flask import send_file
from sqlalchemy import func
from config import db, Config
from models.models import Vehiculo, LineaTransporte, Chofer, municicipio

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.image import Image as XLImage

_MES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

def _register_font():
    try:
        pdfmetrics.registerFont(TTFont("TimesNewRoman", "times.ttf"))
        return "TimesNewRoman"
    except:
        return "Helvetica"

_FONT_NAME = _register_font()
styles = getSampleStyleSheet()
styleN = styles["Normal"]


styleH = styles["Heading1"]
REPORT_LOGO_LEFT = Config.REPORT_LOGO_LEFT
REPORT_LOGO_RIGHT_CARIRUBANA = Config.REPORT_LOGO_RIGHT_CARIRUBANA
REPORT_LOGO_RIGHT_LT = Config.REPORT_LOGO_RIGHT_LT

def _query_vehiculos(municipios=None, combustible=None, grupo=None, max_limit=5000):
    q = db.session.query(Vehiculo).join(LineaTransporte).join(municicipio)
    if municipios:
        if isinstance(municipios, str):
            municipios = [m.strip() for m in municipios.split(",") if m.strip()]
        q = q.filter(municicipio.nombre.in_(municipios))
    if combustible:
        q = q.filter(func.lower(Vehiculo.combustible) == combustible.lower())
    if grupo:
        q = q.filter(func.upper(Vehiculo.grupo) == grupo.upper())
    return q.order_by(LineaTransporte.nombre_organizacion, Vehiculo.placa).limit(max_limit).all()

def _vehiculo_to_row(v):
    linea = getattr(v, "linea", None)
    choferes = getattr(v, "choferes", []) or []
    chofer = choferes[0] if choferes else None
    return {
        "placa": v.placa,
        "marca": v.marca,
        "modelo": v.modelo,
        "capacidad": v.capacidad,
        "litraje": float(v.litraje or 0),
        "sindicato": v.sindicato or "",
        "modalidad": v.modalidad or "",
        "grupo": v.grupo or "",
        "estado": (v.estado or "").lower(),
        "propietario": v.nombre_propietario,
        "cedula_propietario": v.cedula_propietario,
        "linea": linea.nombre_organizacion if linea else "",
        "chofer": chofer.nombre if chofer else "",
        "cedula_chofer": chofer.cedula if chofer else "",
        "combustible": v.combustible or ""
    }

def add_image_right_edge(ws, image_path, row, img_width_px, img_height_px):
    img = XLImage(image_path)
    img.width = img_width_px
    img.height = img_height_px

    col_letter = "O"  # última columna real
    col_idx = column_index_from_string(col_letter) - 1

    # ancho real de la columna en píxeles (aprox)
    col_width = ws.column_dimensions[col_letter].width or 10
    col_width_px = int(col_width * 7)  # conversión Excel → px

    dx = max(col_width_px - img_width_px, 0)

    marker = AnchorMarker(
        col=col_idx,
        colOff=dx * 9525,
        row=row - 1,
        rowOff=0
    )

    size = XDRPositiveSize2D(
        cx=img_width_px * 9525,
        cy=img_height_px * 9525
    )

    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)

def generar_reporte_pdf_bytes(municipios=None, combustible=None, grupo=None, title=None, **kwargs):
    import io
    from datetime import datetime
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, Image
    )
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors

    # =========================
    # FUENTES (TIMES FORZADO)
    # =========================
    FONT = "Times-Roman"
    FONT_BOLD = "Times-Bold"
    FONT_ITALIC = "Times-Italic"
    FONT_BOLD_ITALIC = "Times-BoldItalic"

    ROJO = colors.HexColor("#C60505")
    GRIS = colors.HexColor("#BFBFBF")
    AMARILLO = colors.HexColor("#FFFF00")
    AZUL = colors.HexColor("#B7C4D3")

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    PAGE_WIDTH = landscape(A3)[0]
    AVAILABLE_WIDTH = PAGE_WIDTH - doc.leftMargin - doc.rightMargin

    vehiculos = _query_vehiculos(municipios, combustible, grupo)
    registros = [_vehiculo_to_row(v) for v in vehiculos]

    story = []

    mes = _MES_ES.get(datetime.now().month, "").upper()
    modalidad = "MASIVO (DIESEL)" if any(r.get("combustible") == "Diésel" for r in registros) else "MASIVOS / POR PUESTO / TAXI / MOTO TAXI"

    # =========================
    # LOGOS
    # =========================
    try:
        img_left = Image(REPORT_LOGO_LEFT, width=120, height=100)
    except:
        img_left = Paragraph("", ParagraphStyle("x"))

    if grupo and grupo.upper() == "CARIRUBANA":
        REPORT_LOGO_RIGHT = REPORT_LOGO_RIGHT_CARIRUBANA
        try:
            img_right = Image(REPORT_LOGO_RIGHT, width=100, height=100)
        except:
            img_right = Paragraph("", ParagraphStyle("x"))
    else:
        REPORT_LOGO_RIGHT = REPORT_LOGO_RIGHT_LT
        try:
            img_right = Image(REPORT_LOGO_RIGHT, width=180, height=150)
        except:
            img_right = Paragraph("", ParagraphStyle("x"))


    try:
        img_right = Image(REPORT_LOGO_RIGHT, width=100, height=100)
    except:
        img_right = Paragraph("", ParagraphStyle("x"))

    # =========================
    # ESTILOS
    # =========================
    title_style = ParagraphStyle(
        "title",
        fontName=FONT_BOLD,
        fontSize=20,
        alignment=TA_CENTER
    )

    header_style = ParagraphStyle(
        "header",
        fontName=FONT_BOLD,
        fontSize=18,
        textColor=ROJO,
        alignment=TA_CENTER
    )

    sub_style = ParagraphStyle(
        "sub",
        fontName=FONT_BOLD_ITALIC,
        fontSize=18,
        textColor=ROJO,
        alignment=TA_CENTER
    )

    # =========================
    # ENCABEZADO
    # =========================
    # Estilos diferenciados
    style_black = ParagraphStyle(
        "black",
        fontName=FONT_BOLD,
        fontSize=16,
        alignment=TA_CENTER,
        textColor=colors.black,
        leading=22
    )

    style_red = ParagraphStyle(
        "red",
        fontName=FONT_BOLD,
        fontSize=16,
        alignment=TA_CENTER,
        textColor=ROJO,
        leading=22
    )

    fila_mes_modalidad = Table([
        [Paragraph(mes, style_red), Paragraph(modalidad, style_black)]
    ], colWidths=[(AVAILABLE_WIDTH - 200)/2, (AVAILABLE_WIDTH - 200)/2])

    fila_mes_modalidad.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))

    if grupo and grupo.upper() == "CARIRUBANA":
        bloque_central = [
            Paragraph("INSTITUTO MUNICIPAL DE TRÁNSITO Y TRANSPORTE PÚBLICO DE PASAJEROS DEL MUNICIPIO CARIRUBANA", style_black),
            Paragraph("REGISTRO DE ORGANIZACIONES DE TRANSPORTE PÚBLICO URBANO DEL MUNICIPIO CARIRUBANA", style_red),
            fila_mes_modalidad
        ]
    else:
        bloque_central = [
            Paragraph("REGISTRO DE ORGANIZACIONES DE TRANSPORTE PUBLICO URBANO", style_red),
            Paragraph("MUNICIPIO FALCÓN Y LOS TAQUES", style_red),
            fila_mes_modalidad
        ]
    # Encabezado con logos y bloque central
    encabezado = Table([
        [img_left, bloque_central, img_right]
    ], colWidths=[90, AVAILABLE_WIDTH - 190, 100])

    encabezado.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))

    encabezado.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))

    story.append(encabezado)
    story.append(Spacer(1, 20))


    # =========================
    # TABLA PRINCIPAL
    # =========================
    headers = [
        "N°", "ORGANIZACIÓN", "PROPIETARIO", "C.I",
        "CHOFER", "C.I", "PLACA", "MARCA", "MODELO",
        "PUESTOS", "LITRAJE", "SINDICATO",
        "GRUPO", "MODALIDAD", "FIRMA"
    ]

    data = [headers]

    for i, r in enumerate(registros, 1):
        firma = "INACTIVO - NO FIRMA" if r["estado"] in ("inactivo", "suspendido") else ""
        data.append([
            i, r["linea"], r["propietario"], r["cedula_propietario"],
            r["chofer"], r["cedula_chofer"], r["placa"], r["marca"],
            r["modelo"], r["capacidad"], r["litraje"], r["sindicato"],
            r["grupo"], r["modalidad"], firma
        ])

    raw_widths = [
        30, 140, 140, 65, 140, 65, 75,
        75, 75, 55, 75, 100, 55, 100, 150
    ]

    scale = AVAILABLE_WIDTH / sum(raw_widths)
    col_widths = [w * scale for w in raw_widths]

    row_heights = [42] + [38] * (len(data) - 1)

    tabla = Table(
        data,
        colWidths=col_widths,
        rowHeights=row_heights,
        repeatRows=1
    )

    ts = TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("BACKGROUND", (0, 0), (-1, 0), GRIS),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
    ])

    for idx in range(1, len(data)):
        estado = registros[idx - 1]["estado"]
        if estado == "suspendido":
            ts.add("BACKGROUND", (0, idx), (-1, idx), colors.red)
        elif estado == "inactivo":
            ts.add("BACKGROUND", (0, idx), (-1, idx), AMARILLO)

    tabla.setStyle(ts)
    story.append(tabla)
    story.append(Spacer(1, 30))

    # =========================
    # RESUMEN (CLON EXCEL)
    # =========================
    resumen = {}
    total = 0

    for r in registros:
        if r["estado"] in ("suspendido", "inactivo"):
            continue
        sindicato = r["sindicato"] or "SIN SINDICATO"
        resumen[sindicato] = resumen.get(sindicato, 0) + r["litraje"]
        total += r["litraje"]

    resumen_data = [
        ["", "MUNICIPIO CARIRUBANA", "", "", ""],
        ["N°", "SINDICATOS", "LITROS", "DIA ANTERIOR", "DIFERENCIA"]
    ]

    i = 1
    for s, t in resumen.items():
        resumen_data.append([i, s, t, "", -t])
        i += 1

    resumen_data.append(["", "TOTAL LITRAJE", total, 0, -total])

    col_widths_resumen = [
        AVAILABLE_WIDTH * 0.06,
        AVAILABLE_WIDTH * 0.30,
        AVAILABLE_WIDTH * 0.18,
        AVAILABLE_WIDTH * 0.28,
        AVAILABLE_WIDTH * 0.18
    ]

    resumen_tabla = Table(
        resumen_data,
        colWidths=col_widths_resumen,
        rowHeights=[42, 40] + [36] * (len(resumen_data) - 3) + [40]
    )

    resumen_tabla.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 14),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 1), (-1, -1), 0.6, colors.black),
        ("SPAN", (1, 0), (-2, 0)),
        ("FONTNAME", (1, 0), (1, 0), FONT_BOLD),
        ("BACKGROUND", (2, 2), (2, -1), AMARILLO),
        ("BACKGROUND", (3, 2), (3, -1), GRIS),
        ("BACKGROUND", (4, 2), (4, -1), AZUL),
    ]))

    story.append(resumen_tabla)

    doc.build(story)
    buffer.seek(0)
    return buffer

def generar_reporte_excel_bytes(municipios=None, combustible=None, grupo=None, title=None, **kwargs):
    vehiculos = _query_vehiculos(municipios, combustible, grupo)
    registros = [_vehiculo_to_row(v) for v in vehiculos]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Vehículos"
    ws.sheet_view.showGridLines = False
    mes = _MES_ES.get(datetime.now().month, "").upper()
    NO_BORDER= Border()

    es_diesel = any(r.get("combustible") == "Diésel" for r in registros)

    texto_modalidad = "MASIVO (DIESEL)" if es_diesel else "MASIVOS / POR PUESTO / TAXI / MOTO TAXI"

    # Ajuste de anchos de columna (A y O reservadas para logos; tabla va B..O)
    base_widths = [120, 120, 50, 120, 50, 60, 60,
                   60, 38, 38, 120, 38, 80, 250]
    excel_widths = [round(w * 0.24, 2) for w in base_widths]  # B..O
    # Column A (logo izq) y O (logo der) con ancho fijo para que no se tape
    ws.column_dimensions["A"].width = 10
    for i, w in enumerate(excel_widths, start=2):  # B=2 .. O=15
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.column_dimensions["O"].width = max(ws.column_dimensions["O"].width, 10)

    # Alturas del encabezado (fila 1 más alta para los logos)
    ws.row_dimensions[1].height = 60
    for r in (2, 3, 4):
        ws.row_dimensions[r].height = 25

    # Encabezado
    ws.merge_cells("B1:O1")
    ws.merge_cells("B2:O2")
    ws.merge_cells("B4:D4")
    ws.merge_cells("H4:N4")
    for row in range(1, 5):        # filas 1 a 4
        for col in range(1, 16):   # columnas A a O
            cell = ws.cell(row=row, column=col)
            cell.border = NO_BORDER

    if grupo and grupo.upper() == "CARIRUBANA":
        REPORT_LOGO_RIGHT = REPORT_LOGO_RIGHT_CARIRUBANA
    else:
        REPORT_LOGO_RIGHT = REPORT_LOGO_RIGHT_LT
    
    if grupo and grupo.upper() == "CARIRUBANA":
        ws["B1"].value = "INSTITUTO MUNICIPAL DE TRÁNSITO Y TRANSPORTE PÚBLICO DE PASAJEROS DEL MUNICIPIO CARIRUBANA"
    else:
        ws["B1"].value = "REGISTRO DE ORGANIZACIONES DE TRANSPORTE PUBLICO URBANO"
    ws["B1"].font = Font(name="TimesNewRoman", bold=True, size=20)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    if grupo and grupo.upper() == "CARIRUBANA":
        ws["B2"].value = "REGISTRO DE ORGANIZACIONES DE TRANSPORTE PÚBLICO URBANO DEL MUNICIPIO CARIRUBANA"
    else:
        ws["B2"].value = "MUNICIPIO FALCÓN Y LOS TAQUES"
    ws["B2"].font = Font(name="TimesNewRoman", bold=True, size=18, color="FFC60505")  # rojo oscuro
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["B4"].value = mes
    ws["B4"].font = Font(name="ArialTIME", bold=True, italic=True, size=25, color="FFC60505")
    ws["B4"].alignment = Alignment(horizontal="right", vertical="center")

    ws["H4"].value = texto_modalidad
    ws["H4"].font = Font(name="TimesNewRoman", italic=True, bold=True, size=23) 
    ws["H4"].alignment = Alignment(horizontal="center", vertical="center")

    if REPORT_LOGO_LEFT and os.path.exists(REPORT_LOGO_LEFT):
        img_l = XLImage(REPORT_LOGO_LEFT)
        img_l.width, img_l.height = 200, 180 
        ws.add_image(img_l, "B1")
    else:
        ws["A1"].value = "LOGO"
        ws["A1"].alignment = Alignment(horizontal="right", vertical="center")
        ws["A1"].font = Font(bold=True, size=10)

    if grupo and grupo.upper() == "CARIRUBANA":
        if REPORT_LOGO_RIGHT and os.path.exists(REPORT_LOGO_RIGHT):
            add_image_right_edge(
                ws,
                REPORT_LOGO_RIGHT,
                row=1,
                img_width_px=180,
                img_height_px=180
            )
        else:
            ws["O1"].value = "LOGO"
            ws["O1"].alignment = Alignment(horizontal="left", vertical="center")
            ws["O1"].font = Font(bold=True, size=10)
    else:
        if REPORT_LOGO_RIGHT and os.path.exists(REPORT_LOGO_RIGHT):
            add_image_right_edge(
                ws,
                REPORT_LOGO_RIGHT,
                row=1,
                img_width_px=250,
                img_height_px=180
            )
        else:
            ws["O1"].value = "LOGO"
            ws["O1"].alignment = Alignment(horizontal="left", vertical="center")
            ws["O1"].font = Font(bold=True, size=10)


    # Tabla principal
    start_row = 6
    headers = ["Nº", "ORGANIZACIÓN", "PROPIETARIO - NOMBRE Y APELLIDO", "C.I",
               "CHOFER - NOMBRE Y APELLIDO", "C.I", "PLACA", "MARCA", "MODELO",
               "Nº DE PUESTOS", "LITRAJE", "SINDICATO", "GRUPO", "MODALIDAD", "FIRMA"]

    ws.row_dimensions[start_row].height = 45
    thin = Side(border_style="thin", color="000000")

    for col, h in enumerate(headers, start=1):  # columnas A..O (15)
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = Font(name="TimesNewRoman", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor="BFBFBF")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for i, r in enumerate(registros, start=1):
        row = start_row + i
        ws.row_dimensions[row].height = 45
        firma = "INACTIVO - NO FIRMA" if r["estado"] in ("inactivo", "suspendido") else ""
        values = [i, r["linea"], r["propietario"], r["cedula_propietario"], r["chofer"], r["cedula_chofer"],
                  r["placa"], r["marca"], r["modelo"], r["capacidad"], r["litraje"], r["sindicato"],
                  r["grupo"], r["modalidad"], firma]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="TimesNewRoman", italic=True, bold=True, size=12)
            c.fill = PatternFill("solid", fgColor="F5F5F5")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if r["estado"] == "suspendido":
                c.fill = PatternFill("solid", fgColor="FF0000")
            elif r["estado"] == "inactivo":
                c.fill = PatternFill("solid", fgColor="FFFF00")
            if col in (7, 14):
                c.font = Font(name="TimesNewRoman", size=16, bold=True, italic=True)
            elif col in (4, 6): # C.I.
                c.font = Font(name="TimesNewRoman", size=12, bold=True)
            else:
                c.font = Font(name="TimesNewRoman", size=16, italic=True, bold=True)
            if col == 11:  # LITRAJE
                c.font = Font(name="TimesNewRoman", size=23, italic=True, bold=True)
            elif col == 13:  # GRUPO
                c.font = Font(name="TimesNewRoman", size=23, italic=True, bold=True)

  # =========================
# RESUMEN DE LITRAJE
# =========================
    resumen_start = start_row + len(registros) + 2

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # TÍTULO
    ws.merge_cells(start_row=resumen_start, start_column=1, end_row=resumen_start, end_column=7)
    c = ws.cell(row=resumen_start, column=1, value="RESUMEN DE LITRAJE POR SINDICATO")
    c.font = Font(name="TimesNewRoman", bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ENCABEZADOS
    row_headers = resumen_start + 1

    ws.cell(row=row_headers, column=1, value="N°")
    ws.cell(row=row_headers, column=2, value="SINDICATOS")
    ws.cell(row=row_headers, column=3, value="LITROS")

    # DIA ANTERIOR (4–5)
    ws.merge_cells(start_row=row_headers, start_column=4, end_row=row_headers, end_column=5)
    ws.cell(row=row_headers, column=4, value="DIA ANTERIOR")

    # DIFERENCIA (6–7)
    ws.merge_cells(start_row=row_headers, start_column=6, end_row=row_headers, end_column=7)
    ws.cell(row=row_headers, column=6, value="DIFERENCIA")

    # Estilos encabezados
    for col in (1, 2, 3, 4, 6):
        c = ws.cell(row=row_headers, column=col)
        c.font = Font(name="TimesNewRoman", bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    ws.cell(row=row_headers, column=3).fill = PatternFill("solid", fgColor="FFFF00")
    ws.cell(row=row_headers, column=4).fill = PatternFill("solid", fgColor="BFBFBF")
    ws.cell(row=row_headers, column=6).fill = PatternFill("solid", fgColor="B7C4D3")

    # Cerrar bordes de celdas “fantasma”
    ws.cell(row=row_headers, column=5).border = border
    ws.cell(row=row_headers, column=7).border = border

    # CALCULAR RESUMEN
    resumen = {}
    total_litraje = 0

    for r in registros:
        if r["estado"] in ("suspendido", "inactivo"):
            continue
        sindicato = r["sindicato"] or "SIN SINDICATO"
        litros = r["litraje"]
        resumen[sindicato] = resumen.get(sindicato, 0) + litros
        total_litraje += litros

    # FILAS
    row = row_headers + 1
    contador = 1

    for sindicato, litros in resumen.items():
        ws.cell(row=row, column=1, value=contador)
        ws.cell(row=row, column=2, value=sindicato)
        ws.cell(row=row, column=3, value=litros)

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws.cell(row=row, column=4, value="")

        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6, value=-litros)

        for col in (1, 2, 3, 4, 6):
            c = ws.cell(row=row, column=col)
            c.font = Font(name="TimesNewRoman", size=14)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor="FFFF00")
        ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="BFBFBF")
        ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor="B7C4D3")

        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=7).border = border

        ws.row_dimensions[row].height = 28
        contador += 1
        row += 1

    # TOTAL GENERAL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="TOTAL LITRAJE")

    ws.cell(row=row, column=3, value=total_litraje)

    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    ws.cell(row=row, column=4, value=0)

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    ws.cell(row=row, column=6, value=-total_litraje)

    for col in (1, 3, 4, 6):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="TimesNewRoman", bold=True, size=16)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor="FFFF00")
    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="BFBFBF")
    ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor="B7C4D3")

    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=7).border = border
    ws.row_dimensions[row].height = 32



    ws.print_options.fitToPage = True
    ws.print_options.fitToWidth = 1
    ws.print_options.fitToHeight = 1
    ws.page_setup.orientation = 'landscape'  

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def generar_reporte_pdf_response(**params):
    buf = generar_reporte_pdf_bytes(**params)
    filename = f"reporte_vehiculos_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


def generar_reporte_excel_response(**params):
    buf = generar_reporte_excel_bytes(**params)
    filename = f"reporte_vehiculos_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

